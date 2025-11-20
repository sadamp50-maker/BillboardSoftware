# app.py
# Billboard Rental Management — Streamlit application
# Save this file as `app.py` in your project folder (e.g. C:\Users\K.C\Desktop\BillboardSoftware)

import streamlit as st
import pandas as pd
import os
from datetime import datetime, timedelta
from pathlib import Path
from openpyxl import load_workbook
import sqlite3

# ---------------------------
# CONFIG: paths
# ---------------------------
# Prefer a local copy in project folder (Billboard.xlsm). If not found, fall back to the uploaded path.
LOCAL_COPY = "Billboard.xlsm"
UPLOADED_COPY = "/mnt/data/Billboard.xlsm"  # path where your uploaded workbook is stored in this environment

if os.path.exists(LOCAL_COPY):
    EXCEL_PATH = LOCAL_COPY
elif os.path.exists(UPLOADED_COPY):
    EXCEL_PATH = UPLOADED_COPY
else:
    EXCEL_PATH = None

# SQLite (optional) — local file
DB_PATH = "billboards.db"

# ---------------------------
# Helpers to read / write sheets
# ---------------------------

def read_sheets():
    """Return (dashboard_df, summary_df, saved_df, sheet_names)
    If EXCEL_PATH is None, fallback to CSV simple storage.
    """
    if EXCEL_PATH and os.path.exists(EXCEL_PATH):
        xls = pd.ExcelFile(EXCEL_PATH)
        sheet_names = xls.sheet_names
        # try to find best-match names
        dash_name = next((s for s in sheet_names if 'Dashboard' in s or 'Billboard' in s or 'Rentals' in s), sheet_names[0])
        summary_name = next((s for s in sheet_names if 'Summary' in s), sheet_names[1] if len(sheet_names)>1 else sheet_names[0])
        saved_name = next((s for s in sheet_names if 'Saved' in s or 'Archive' in s), sheet_names[-1])

        dashboard = pd.read_excel(xls, sheet_name=dash_name, dtype=str).fillna("")
        summary = pd.read_excel(xls, sheet_name=summary_name, dtype=str).fillna("")
        saved = pd.read_excel(xls, sheet_name=saved_name, dtype=str).fillna("")

        # normalize obvious date columns
        for df in (dashboard, saved):
            for c in df.columns:
                if any(k in c.lower() for k in ['date','from','to','start','end']):
                    try:
                        df[c] = pd.to_datetime(df[c], errors='coerce')
                    except Exception:
                        pass
        return dashboard, summary, saved, (dash_name, summary_name, saved_name)

    # fallback
    if not os.path.exists('saved_data.csv'):
        pd.DataFrame(columns=["Billboard Number","Client Name","Start Date","End Date","Rent Amount","Remarks"]).to_csv('saved_data.csv', index=False)
    saved = pd.read_csv('saved_data.csv')
    dashboard = pd.DataFrame(columns=saved.columns)
    summary = pd.DataFrame({"Total Boards":[0]})
    return dashboard, summary, saved, ("Dashboard","Summary Dashboard","SavedData")


def save_to_xlsm(dashboard_df, summary_df, saved_df, sheet_names):
    """Save dataframes back into EXCEL_PATH. If EXCEL_PATH is None, write to CSV files."""
    if not EXCEL_PATH:
        saved_df.to_csv('saved_data.csv', index=False)
        dashboard_df.to_csv('dashboard_data.csv', index=False)
        return

    wb = load_workbook(EXCEL_PATH, keep_vba=True)
    dash_name, summary_name, saved_name = sheet_names

    def replace_sheet(wsname, df):
        if wsname in wb.sheetnames:
            ws = wb[wsname]
            # clear existing cells (simple approach)
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
            # write header
            for j, col in enumerate(df.columns, start=1):
                ws.cell(row=1, column=j).value = col
            # write rows
            for i, row in enumerate(df.itertuples(index=False), start=2):
                for j, value in enumerate(row, start=1):
                    if pd.isna(value):
                        v = None
                    else:
                        v = value
                        if isinstance(v, pd.Timestamp):
                            v = v.strftime('%Y-%m-%d')
                    ws.cell(row=i, column=j).value = v
        else:
            ws = wb.create_sheet(wsname)
            for j, col in enumerate(df.columns, start=1):
                ws.cell(row=1, column=j).value = col
            for i, row in enumerate(df.itertuples(index=False), start=2):
                for j, value in enumerate(row, start=1):
                    v = None if pd.isna(value) else value
                    if isinstance(v, pd.Timestamp):
                        v = v.strftime('%Y-%m-%d')
                    ws.cell(row=i, column=j).value = v

    replace_sheet(dash_name, dashboard_df)
    replace_sheet(summary_name, summary_df)
    replace_sheet(saved_name, saved_df)

    wb.save(EXCEL_PATH)

# ---------------------------
# Start — load data
# ---------------------------

dashboard_df, summary_df, saved_df, sheet_names = read_sheets()

# attempt to coerce numeric columns
for c in saved_df.columns:
    if any(k in c.lower() for k in ['rent','amount','price']):
        saved_df[c] = pd.to_numeric(saved_df[c], errors='coerce').fillna(0)

# ---------------------------
# SQLite helper (optional)
# ---------------------------

def ensure_db():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    return conn

# ==========================
# Streamlit UI
# ==========================

st.set_page_config(page_title="Billboard Manager", layout="wide")
st.title("📊 Billboard Rental Management System")

st.sidebar.header("Settings")
use_sql = st.sidebar.checkbox("Mirror to SQLite DB (optional)", value=False)
auto_save = st.sidebar.checkbox("Auto-save to Excel/XLSM", value=True)
alert_days = st.sidebar.number_input("Alert before expiry (days)", min_value=0, max_value=365, value=7)

if use_sql:
    conn = ensure_db()
else:
    conn = None

menu = st.sidebar.radio("View", ["Dashboard","Summary Dashboard","Saved Data","Admin"])

# helper: compute status

def compute_status(end_val):
    try:
        if pd.isna(end_val) or end_val == "":
            return 'Available'
        end_dt = pd.to_datetime(end_val)
    except Exception:
        return 'Unknown'
    today = pd.Timestamp(datetime.today().date())
    if end_dt < today:
        return 'Expired'
    if end_dt <= today + pd.Timedelta(days=alert_days):
        return 'Expiring Soon'
    return 'Booked'

# -------------------------
# Dashboard view
# -------------------------
if menu == 'Dashboard':
    st.header('📋 Dashboard — Edit / Add')
    st.markdown('نیچے Dashboard شیٹ editable grid کے طور پر دکھائی گئی ہے۔')

    try:
        edited = st.data_editor(dashboard_df, num_rows='dynamic', use_container_width=True)
    except Exception:
        st.write(dashboard_df)
        edited = dashboard_df

    st.markdown('### Archive a row (move to SavedData)')
    col1, col2 = st.columns([3,1])
    with col1:
        st.write('Select row index to archive (0-based):')
        idx = st.number_input('Row index', min_value=0, max_value=max(0, len(edited)-1), value=0, step=1)
    with col2:
        if st.button('Archive selected row'):
            row = edited.iloc[idx]
            saved_df = pd.concat([saved_df, pd.DataFrame([row])], ignore_index=True)
            edited.iloc[idx] = [''] * len(edited.columns)
            dashboard_df = edited
            if auto_save:
                save_to_xlsm(dashboard_df, summary_df, saved_df, sheet_names)
            st.success(f'Row {idx} archived')
            st.experimental_rerun()

    st.markdown('---')
    st.subheader('➕ Quick Add')
    with st.form('add'):
        cols = st.columns(4)
        bnum = cols[0].text_input('Billboard Number')
        client = cols[1].text_input('Client Name')
        start = cols[2].date_input('Start Date', value=datetime.today())
        end = cols[3].date_input('End Date', value=datetime.today()+timedelta(days=30))
        rent = st.number_input('Rent Amount', min_value=0.0, value=0.0)
        remarks = st.text_input('Remarks')
        submit = st.form_submit_button('Add Entry')
        if submit:
            new = {c: '' for c in edited.columns}
            for c in edited.columns:
                lc = c.lower()
                if 'bill' in lc or 'number' in lc:
                    new[c] = bnum
                elif 'client' in lc or 'name' in lc:
                    new[c] = client
                elif 'start' in lc or 'from' in lc:
                    new[c] = start
                elif 'end' in lc or 'to' in lc:
                    new[c] = end
                elif 'rent' in lc or 'amount' in lc or 'price' in lc:
                    new[c] = rent
                elif 'remark' in lc or 'note' in lc:
                    new[c] = remarks
            dashboard_df = pd.concat([edited, pd.DataFrame([new])], ignore_index=True)
            if auto_save:
                save_to_xlsm(dashboard_df, summary_df, saved_df, sheet_names)
            st.success('Entry added')
            st.experimental_rerun()

# -------------------------
# Summary view
# -------------------------
elif menu == 'Summary Dashboard':
    st.header('📈 Summary Dashboard')
    df = saved_df.copy()
    amt_col = next((c for c in df.columns if any(k in c.lower() for k in ['rent','amount','price'])), None)
    total_rent = df[amt_col].astype(float).sum() if amt_col is not None else 0
    st.metric('Total Saved Records', df.shape[0])
    st.metric('Total Rent Collected', total_rent)
    st.metric('Unique Clients', df.iloc[:,1].nunique() if df.shape[1]>1 else 0)

    st.markdown('### Due / Expiry Alerts')
    if df.shape[0]>0:
        end_col = next((c for c in df.columns if 'end' in c.lower()), None)
        if end_col:
            df['_end_dt'] = pd.to_datetime(df[end_col], errors='coerce')
            today = pd.Timestamp(datetime.today().date())
            expired = df[df['_end_dt'] < today]
            soon = df[(df['_end_dt'] >= today) & (df['_end_dt'] <= today + pd.Timedelta(days=alert_days))]
            if not expired.empty:
                st.warning(f"{len(expired)} records expired")
                st.dataframe(expired.drop(columns=['_end_dt']))
            if not soon.empty:
                st.info(f"{len(soon)} records expiring within {alert_days} days")
                st.dataframe(soon.drop(columns=['_end_dt']))
    st.markdown('### Recent Saved')
    st.dataframe(df.tail(20))

# -------------------------
# Saved Data view
# -------------------------
elif menu == 'Saved Data':
    st.header('📁 Saved Data (Archive)')
    st.dataframe(saved_df)
    c1, c2 = st.columns(2)
    with c1:
        if st.button('Export CSV'):
            saved_df.to_csv('SavedData_export.csv', index=False)
            st.success('SavedData_export.csv created')
    with c2:
        if st.button('Clear All'):
            if st.confirm('Are you sure?'):
                saved_df = pd.DataFrame(columns=saved_df.columns)
                if auto_save:
                    save_to_xlsm(dashboard_df, summary_df, saved_df, sheet_names)
                st.success('Cleared')
                st.experimental_rerun()

# -------------------------
# Admin
# -------------------------
elif menu == 'Admin':
    st.header('⚙️ Admin')
    if st.button('Save to Excel/XLSM'):
        save_to_xlsm(dashboard_df, summary_df, saved_df, sheet_names)
        st.success('Saved to Excel/XLSM (backup recommended)')
    if st.button('Reload from Excel/XLSM'):
        dashboard_df, summary_df, saved_df, sheet_names = read_sheets()
        st.experimental_rerun()
    if st.button('Sync to SQLite'):
        try:
            conn = ensure_db()
            saved_df.to_sql('saveddata', conn, if_exists='replace', index=False)
            st.success('Synced to SQLite')
        except Exception as e:
            st.error('DB sync failed: ' + str(e))

st.sidebar.markdown('---')
st.sidebar.caption('Use Admin to save back to your XLSM. Keep a backup before writing if your workbook has macros.')
